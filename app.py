import os
import re
import hmac
from functools import wraps
import pandas as pd
from datetime import datetime, timedelta
from dotenv import load_dotenv
from flask import (
    Flask, render_template, request, send_file, redirect, url_for,
    flash, jsonify, session, abort,
)
from werkzeug.utils import secure_filename

load_dotenv()

import db  # noqa: E402  (must come after load_dotenv)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY") or "dev-insecure-key-change-me"
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50MB max upload

# Session cookie hardening. SECURE is gated on an env flag so local HTTP dev
# still works; set SESSION_COOKIE_SECURE=1 (the default) in production behind HTTPS.
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = (
    os.environ.get("SESSION_COOKIE_SECURE", "1").lower() in ("1", "true", "yes")
)

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")


def require_admin(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("admin_login", next=request.path))
        return view(*args, **kwargs)
    return wrapper


@app.context_processor
def inject_admin_flag():
    return {"is_admin": bool(session.get("is_admin"))}

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
QC_CODES_FILE = os.path.join(BASE_DIR, "RMA_Codes.csv") 
RMA_GUIDELINES_FILE = os.path.join(BASE_DIR, "Vendor RMA Guidelines.xlsx")

# Local Dropbox file for incoming invoices (.csv or .xlsx both supported)
INVOICES_FILE = (
    "/mnt/c/Users/Mannapov/Mannapov LLC Dropbox/Daniel Trevino/Analytics/Google Sheets/Incoming Invoices/IncomingInvoices.csv"
)

# === QC codes that are considered "pass" (not a problem) ===
# LCD-L1 is treated as pass — L1 burns are not sent for RMA.
QC_PASS_CODES = {"M00", "LCD-L0", "LCD-L1"}
# HYLA-only: L2 burns are too subtle to prove in a photo/video, so HYLA
# does not accept them. Only L3-L5 burns are claimable for HYLA.
HYLA_EXTRA_PASS_CODES = {"LCD-L2"}
UNLOCKED_CARRIERS = {"Unlocked", "International (Unlocked)", "Wi-Fi Only", "WiFi", ""}

ICE_COLS = [
    "WID", "Vendor", "Vendor Invoice #", "Vendor Condition", "Vendor Description",
    "Vendor Model",
    "Cost Amount", "Condition", "IMEI",
    "Brand", "Model", "Carrier", "Carrier GSX", "QC Error Code",
    "Physical QC LCD",
    "Physical QC Board", "Inventory", "Inventory Bin #",
    "Inventory Location #", "Inventory Notes (Out Reason)",
    "Inventory By", "Inventory Date",
    "Missing/Extra",
    "Project #", "Notes",
    "Tech Notes",              # HYLA: Battery Door Replacement / Lifted Battery Door = non-returnable
    "Pallet #", "Vendor Notes",  # HYLA RMA form: PO/Order# and Warehouse
]

# ---------------------------------------------------------------------------
# Map QC error codes -> RMA guideline category
# These categories match the row labels in Vendor RMA Guidelines.xlsx
# ---------------------------------------------------------------------------
QC_CODE_TO_RMA_CATEGORY = {
    # Camera failures (functional)
    "M11": "Camera Fail",          # Back Camera
    "M18": "Camera Fail",          # Front Camera
    # Button failures
    "M05": "Buttons Fail",         # Power Button
    "M13": "Buttons Fail",         # Volume Up/Down
    "M14": "Buttons Fail",         # Home Button
    "M21": "Buttons Fail",         # Menu Keys
    "M23": "Buttons Fail",         # Ringer On/Off
    # Touchscreen failures
    "M03": "Touchscreen Fail",     # Touchscreen
    "M28": "Touchscreen Fail",     # 3D Touch
    # Chargeport failures
    "M04": "Chargeport Fail",      # Charge Port
    "M26": "Chargeport Fail",      # Wireless Charging
    # Parts Message
    "R7": "Parts Message",         # Flex Cable Torn
    # Face ID / Touch ID
    "M17": "Face ID Fail",         # Touch ID / Face ID
    # Speaker / Mic failures
    "M08": "Speaker Fail",         # Microphone
    "M12": "Speaker Fail",         # Loud Speaker
    "M19": "Speaker Fail",         # Earpiece Speaker
    "M30": "Speaker Fail",         # Audio Jack
    "FUNC P01": "Speaker Fail",    # Loud Speaker Low Volume
    "FUNC P02": "Speaker Fail",    # Front Mic Low Volume
    "FUNC P03": "Speaker Fail",    # Bottom Mic Low Volume
    # Wifi / Bluetooth
    "M20": "Wifi/Bluetooth Fail",  # WiFi/Bluetooth
    "M27": "Wifi/Bluetooth Fail",  # NFC (connectivity)
    # Cracked Screen
    "M02": "Cracked Screen",       # LCD cosmetic issue
    "R6": "Cracked Back",          # Bent device
    "M34": "Cracked Back",         # Flooded (Polishing)
    # Camera lens (cosmetic crack, not functional camera failure)
    "PHY-P12": "Cracked Camera",   # Camera Lens Replaced
    # Burns (L1 is treated as pass — see QC_PASS_CODES)
    "LCD-L2": "Burns (L2-L5)",     # Very Light Burn Images
    "LCD-L3": "Burns (L2-L5)",     # Medium Burn Images
    "LCD-L4": "Burns (L2-L5)",     # Heavy Burn Image
    "LCD-L5": "Burns (L2-L5)",     # Extreme Burn Image
    # LCD failure codes → Cracked Screen (major display damage / cracked glass)
    "LCD-F01": "Cracked Screen",   # Visually Bad
    "LCD-F03": "Bleeding LCD",     # Bleeding LCD (LCD bleed — NOT a cracked screen)
    "LCD-F05": "Cracked Screen",   # Water / 3D Damage / Large White Areas
    "LCD-F06": "Cracked Screen",   # Flickering / Flashing Only
    "LCD-F10": "Cracked Screen",   # Heavy Light Leakage
    "LCD-F11": "Cracked Screen",   # Other (Manager list problem)
    "LCD-F12": "Cracked Screen",   # LCD Main
    "LCD-F13": "Cracked Screen",   # LCD Front
    "LCD-F14": "Cracked Screen",   # LCD Main/Front
    # LCD codes → Missing Pixels (pixel-level issues)
    "LCD-F04": "Missing Pixels",   # Line(s)
    "LCD-F07": "Missing Pixels",   # Heavy Discoloration
    "LCD-P01": "Missing Pixels",   # Other (Manager list flaw)
    "LCD-P03": "Missing Pixels",   # Minor Discoloration
    "LCD-P04": "Missing Pixels",   # Light Leakage
    # LCD codes → Display Spots (spot/pressure-point/spec issues)
    "R2":      "Display Spots",    # Spots
    "LCD-F08": "Display Spots",    # Multiple / Medium / Large Pressure Point(s)
    "LCD-F09": "Display Spots",    # Heavy Brown Spots
    "LCD-P02": "Display Spots",    # Minor Light Spots/Areas
    "LCD-PP1": "Display Spots",    # Pen-Like Spec
    "LCD-PP2": "Display Spots",    # Microscopic Pressure Point
    "LCD-PP3": "Display Spots",    # Small Pressure Point
    # Physical codes → Cracked Screen (front/screen damage)
    "PHY-F04": "Cracked Screen",   # Excessive Lifted Screen
    "PHY-P03": "Cracked Screen",   # Minor Lifted Screen
    "PHY-P14": "Cracked Screen",   # Front Glass Polishing
    # Physical codes → Cracked Back (back/body damage)
    "PHY-F01": "Cracked Back",     # Heavy Delamination
    "PHY-F02": "Cracked Back",     # Excessive Wear/Usage
    "PHY-F03": "Cracked Back",     # Missing Cover
    "PHY-F05": "Cracked Back",     # Bent Device
    "PHY-F09": "Cracked Back",     # Other Fail
    "PHY-P01": "Cracked Back",     # Minor Delamination
    "PHY-P02": "Cracked Back",     # Minor Wear/Usage
    "PHY-P09": "Cracked Back",     # Other Pass
    "PHY-P11": "Cracked Back",     # Battery Door Replaced
    "PHY-P13": "Cracked Back",     # Chrome Side Polishing
    "PHY-P15": "Cracked Back",     # Glass Polishing Outsourced
    "PHY-P17": "Cracked Back",     # Stylus Replaced
    # ID / MDM Lock
    "M31": "ID/MDM Lock",          # Activation Lock
    # Not Cleared / Blacklisted
    "M32": "Not Cleared",          # Blacklisted
    "M33": "Not Cleared",          # Jailbroken
}

# Functional (non-cosmetic, non-lock) categories — used to decide whether an
# unmapped QC code is relevant for a given vendor.
QC_FUNCTIONAL_CATEGORIES = {
    "Camera Fail", "Buttons Fail", "Touchscreen Fail", "Chargeport Fail",
    "Parts Message", "Face ID Fail", "Speaker Fail", "Wifi/Bluetooth Fail",
}

# Condition column keywords -> RMA category
# Cosmetic fails are detected ONLY from the Condition column text, not from QC codes.
# Order matters — first match wins. Specific sub-type keywords MUST come before the
# broad "grade d" / "grade f" catchalls; otherwise "Grade D - Back Glass Only" would
# match "grade d" first and get labeled Cracked Screen instead of Cracked Back.
# NOTE: Display burns and display spots are intentionally NOT detected here —
# they come from the structured QC columns (QC Error Code + Physical QC LCD).
CONDITION_TO_RMA_CATEGORY = {
    # --- New - Activated ---
    "new - activated":      "New - Activated",
    "newly activated":      "New - Activated",
    # --- Cracked Camera (specific — must come before "grade d") ---
    "camera lens":          "Cracked Camera",  # Grade D (Camera Lens Only)
    # --- Cracked Back (specific — must come before "grade d") ---
    "back glass":           "Cracked Back",    # Grade D (Back Glass Only)
    "back cover":           "Cracked Back",    # Grade D (Back Cover dent/bent)
    "bent":                 "Cracked Back",    # Bent device
    # --- Cracked Screen (specific) ---
    "bad lcd":              "Cracked Screen",  # Grade F (Bad LCD)
    "display failed":       "Cracked Screen",  # Grade F (Display Failed Only)
    "display imperfection": "Cracked Screen",  # Grade AB (Display Imperfection)
    "front glass":          "Cracked Screen",  # Grade D (Front Glass Only)
    "lifted screen":        "Cracked Screen",  # Lifted screen (vendor groups w/ Cracked Screen)
    "ber/scrap":            "Cracked Screen",  # BER / Scrap (severe damage)
    # --- Other ---
    "parts message":        "Parts Message",
    "not cleared":          "Not Cleared",
    "id lock":              "ID/MDM Lock",     # "ID Lock" or "ID Locked"
    # --- Generic grade fallbacks (last — only if no specific keyword matched) ---
    "grade f":              "Cracked Screen",  # Grade F generic — usually bad screen
    "grade d":              "Cracked Screen",  # Grade D generic — most common is front glass
}

# ---------------------------------------------------------------------------
# Load per-grade RMA rules from Vendor RMA Guidelines.xlsx
# Produces: { "HYLA TPS": { "A+": { "Camera Fail": "YES"|"NO"|"NOT SURE", ... }, ... }, ... }
# ---------------------------------------------------------------------------
def load_rma_grade_rules():
    if not os.path.exists(RMA_GUIDELINES_FILE):
        return {}
    df = pd.read_excel(RMA_GUIDELINES_FILE, sheet_name="Vendor RMA Guidelines", header=None)

    # Spreadsheet row index -> category name (skips header/section-label rows)
    # Layout: row 0=VENDOR, row 1=GRADE, rows 2-5=metadata, rows 6+=categories
    CATEGORY_ROWS = {
        6:  "Carrier Locked",
        7:  "Not Cleared",
        8:  "New - Activated",
        9:  "ID/MDM Lock",
        10: "Low Battery",
        # row 11 = "COSMETIC FAIL:" section header — skip
        12: "Cracked Screen",
        13: "Bleeding LCD",
        14: "Cracked Back",
        15: "Cracked Camera",
        16: "Burns (L2-L5)",
        17: "Missing Pixels",
        18: "Display Spots",
        # row 19 = "FUNCTION FAIL:" section header — skip
        20: "Camera Fail",
        21: "Buttons Fail",
        22: "Touchscreen Fail",
        23: "Chargeport Fail",
        24: "Parts Message",
        25: "Face ID Fail",
        26: "Speaker Fail",
        27: "Wifi/Bluetooth Fail",
    }

    rules = {}
    for col in range(1, df.shape[1]):
        vendor = str(df.iloc[0, col]).strip()
        grade  = str(df.iloc[1, col]).strip()
        if vendor in ("nan", "") or grade in ("nan", ""):
            continue
        grade_data = {}
        for row_idx, category in CATEGORY_ROWS.items():
            if row_idx >= df.shape[0]:
                grade_data[category] = "NO"
                continue
            raw = str(df.iloc[row_idx, col]).strip().upper()
            if raw.startswith("YES"):
                grade_data[category] = "YES"
            elif "NOT SURE" in raw or "NEED TO TEST" in raw or "ASKING JASON" in raw:
                grade_data[category] = "NOT SURE"
            else:
                grade_data[category] = "NO"
        rules.setdefault(vendor, {})[grade] = grade_data
    return rules

RMA_GRADE_RULES = load_rma_grade_rules()

# HYLA no longer accepts speaker fails — force Speaker Fail to NO for all HYLA grades,
# regardless of what the Vendor RMA Guidelines spreadsheet says.
for _hyla_vendor in ("HYLA TPS", "HYLA DLS"):
    for _grade_data in RMA_GRADE_RULES.get(_hyla_vendor, {}).values():
        _grade_data["Speaker Fail"] = "NO"


def _lcd_state_from_condition(cond):
    """Extract an LCD assertion from the Condition column.
    Returns ('burn', level) | ('clean',) | ('spots',) | ('imp',) | ('bad',) | None.
    'clean' = explicit L0/L1 declaration. None = no LCD-related text.
    """
    s = str(cond or "").lower()
    levels = [int(x) for x in re.findall(r'display burn l(\d)', s)]
    if levels:
        peak = max(levels)
        return ("burn", peak) if peak >= 2 else ("clean",)
    if "display spots" in s:
        return ("spots",)
    if "display imperfection" in s:
        return ("imp",)
    if "display failed" in s or "bad lcd" in s:
        return ("bad",)
    return None


def _lcd_state_from_physical(phys):
    """Extract an LCD assertion from the Physical QC LCD column.
    Returns ('burn', level) | ('clean',) | ('bleed',) | ('imp',) | ('spots',) | None.
    """
    s = str(phys or "").upper().strip()
    if not s:
        return None
    levels = [int(x) for x in re.findall(r'\bL([0-5])\b', s)]
    if levels:
        peak = max(levels)
        return ("burn", peak) if peak >= 2 else ("clean",)
    if "BLEED" in s:
        return ("bleed",)
    if "IMP" in s:
        return ("imp",)
    if "SPOT" in s:
        return ("spots",)
    return None


def _lcd_state_from_qc(qc_codes_str):
    """Extract an LCD assertion from the QC Error Code column.
    Priority: burn > bad > bleed > spots > clean. LCD-L0/L1 alone => 'clean'.
    """
    s = str(qc_codes_str or "").upper()
    levels = [int(x) for x in re.findall(r'LCD-L([0-5])', s)]
    burn_peak = max(levels) if levels else None
    if burn_peak is not None and burn_peak >= 2:
        return ("burn", burn_peak)
    if re.search(r'LCD-(?:F01|F05|F06|F10|F11|F12|F13|F14)\b', s):
        return ("bad",)
    if "LCD-F03" in s:
        return ("bleed",)
    if re.search(r'LCD-(?:P02|PP1|PP2|PP3|F08|F09)\b|(?<!\w)R2(?!\w)', s):
        return ("spots",)
    if burn_peak is not None:  # LCD-L0/L1 with no other LCD claim
        return ("clean",)
    return None


def lcd_mismatch(cond, qc_codes, phys):
    """True if Condition / QC Error Code / Physical QC LCD disagree about LCD.
    Any two non-None claims that differ count as a mismatch.
    """
    states = [s for s in (
        _lcd_state_from_condition(cond),
        _lcd_state_from_qc(qc_codes),
        _lcd_state_from_physical(phys),
    ) if s is not None]
    if len(states) < 2:
        return False
    first = states[0]
    return any(s != first for s in states[1:])


def parse_threshold(raw):
    """Parse a threshold string from RMA guidelines into a percentage float.
    Returns:
        float >= 0: the threshold percentage (0 means no threshold, always eligible)
        None: threshold is NO / unknown / not applicable
    """
    s = str(raw).strip()
    upper = s.upper()
    if upper in ("NO", "NAN", ""):
        return None
    if "NEED TO TEST" in upper or "ASKING" in upper:
        return None
    if "NO THRESHOLD" in upper or "EVERY DEVICE" in upper:
        return 0.0
    match = re.search(r'(\d+(?:\.\d+)?)\s*%', s)
    if match:
        return float(match.group(1))
    return None


def load_rma_thresholds():
    """Load threshold percentages per vendor from RMA guidelines row 3.
    Returns { vendor_name: float_or_None }
    """
    if not os.path.exists(RMA_GUIDELINES_FILE):
        return {}
    df = pd.read_excel(RMA_GUIDELINES_FILE, sheet_name="Vendor RMA Guidelines", header=None)
    if df.shape[0] < 4:
        return {}
    thresholds = {}
    for col in range(1, df.shape[1]):
        vendor = str(df.iloc[0, col]).strip()
        if vendor in ("nan", ""):
            continue
        if vendor not in thresholds:
            thresholds[vendor] = parse_threshold(df.iloc[3, col])
    return thresholds


# Map vendor config keys to their name in the RMA guidelines spreadsheet
VENDOR_KEY_TO_RMA_NAME = {
    "hyla": "HYLA TPS",
    "hyla_tps": "HYLA TPS",
    "hyla_dls": "HYLA DLS",
    "att": "AT&T",
    "sprint": "Sprint",
    "verizon": "Verizon",
    "superior": "Superior (Apple)",
    "touchstone": "Touchstone",
    "clover": "Clover",
}


def get_grade_rules(vendor_name, vendor_condition, fallback_returnable, fallback_not_sure, rma_grade_rules=None):
    """
    Extract the grade from Vendor Condition (e.g. "TPS A+" → "A+", "DLS B+" → "B+")
    and return (returnable, not_sure) sets from rma_grade_rules.
    Falls back to the static vendor sets for vendors not in the guidelines.
    Returns (None, None) if the vendor uses grade rules but the grade isn't listed
    (meaning the device should be skipped entirely).
    """
    rules = rma_grade_rules if rma_grade_rules is not None else RMA_GRADE_RULES
    vendor_rules = rules.get(vendor_name)
    if not vendor_rules:
        # Vendor has no grade-level rules → use static fallback (e.g. AT&T, Sprint)
        return fallback_returnable, fallback_not_sure

    cond = str(vendor_condition or "").strip()
    parts = cond.split()
    grade = parts[-1] if parts else ""

    grade_data = vendor_rules.get(grade)
    if grade_data is None and cond:
        # Robust fallback for carrier-auction grades (e.g. Sprint/T-Mobile
        # "A+", "B-", "C+", "CPO", "DBR") where the grade may not be the last
        # whitespace token or may differ in case. Match a known grade as a
        # whole word, most-specific first so "A+" beats "A" and "B-" beats "B".
        # The "+"/"-" suffix is treated as part of the token so plain "A" is
        # never read out of "A+". Non-standard grades (CPO, DBR, DWF, N, blank)
        # have no column and stay unmatched → device is excluded.
        by_upper = {g.upper(): g for g in vendor_rules}
        grade_data = vendor_rules.get(by_upper.get(grade.upper(), ""))
        if grade_data is None:
            cond_up = cond.upper()
            for g in sorted(vendor_rules, key=len, reverse=True):
                if re.search(rf"(?<![\w+\-]){re.escape(g.upper())}(?![\w+\-])", cond_up):
                    grade_data = vendor_rules[g]
                    break
    if grade_data is None:
        # Grade not in guidelines → suppress device entirely
        return None, None

    returnable = {cat for cat, val in grade_data.items() if val == "YES"}
    not_sure   = {cat for cat, val in grade_data.items() if val == "NOT SURE"}
    return returnable, not_sure


# ---------------------------------------------------------------------------
# Vendor definitions — keyed by a short slug
# `invoice_pattern` is a regex matched against the Vendor column in IncomingInvoices.csv
# `returnable` / `not_sure` are fallbacks for vendors without per-grade rules.
# For HYLA TPS/DLS the per-grade rules from RMA_GRADE_RULES take precedence.
# `grade_prefix_vendors` maps Vendor Condition prefix → RMA rules vendor name (combined vendors).
# ---------------------------------------------------------------------------
VENDORS = {
    "hyla": {
        "name": "HYLA (TPS + DLS)",
        "description": "Combined HYLA TPS and DLS inventory. TPS-graded devices use TPS rules; DLS-graded devices use DLS rules.",
        "invoice_pattern": r"hyla",
        "grade_prefix_vendors": {"TPS": "HYLA TPS", "DLS": "HYLA DLS"},
        "lookback_days": 7,
        "returnable": {
            "ID/MDM Lock",
            "Camera Fail", "Buttons Fail", "Touchscreen Fail",
            "Chargeport Fail", "Parts Message", "Face ID Fail",
            "Cracked Screen", "Bleeding LCD", "Cracked Back", "Cracked Camera", "Burns (L2-L5)",
        },
        "not_sure": {"Carrier Locked", "Not Cleared"},
    },
    "hyla_tps": {
        "name": "HYLA TPS",
        "description": "Source phones from HYLA TPS inventory. Return via HYLA TPS RMA portal — rules vary by grade (see grade table below).",
        "invoice_pattern": r"hyla",
        "vendor_grade_prefix": "TPS",
        "exclude_grade_prefix": "DLS",   # exclude explicitly DLS-labeled devices
        "lookback_days": 7,
        "include_in_all": False,         # covered by combined "hyla" entry

        # Fallback only — actual rules come from RMA_GRADE_RULES per device grade
        "returnable": {
            "ID/MDM Lock",
            "Camera Fail", "Buttons Fail", "Touchscreen Fail",
            "Chargeport Fail", "Parts Message", "Face ID Fail",
            "Cracked Screen", "Bleeding LCD", "Cracked Back", "Cracked Camera", "Burns (L2-L5)",
        },
        "not_sure": {"Carrier Locked", "Not Cleared"},
    },
    "hyla_dls": {
        "name": "HYLA DLS",
        "description": "Source phones from HYLA DLS inventory. Return via HYLA DLS RMA portal — rules vary by grade (see grade table below).",
        "invoice_pattern": r"hyla",
        "vendor_grade_prefix": "DLS",
        "exclude_grade_prefix": "TPS",   # exclude explicitly TPS-labeled devices
        "lookback_days": 7,
        "include_in_all": False,         # covered by combined "hyla" entry

        # Fallback only — actual rules come from RMA_GRADE_RULES per device grade
        "returnable": {
            "ID/MDM Lock",
            "Camera Fail", "Buttons Fail", "Touchscreen Fail",
            "Chargeport Fail", "Parts Message", "Face ID Fail",
            "Cracked Screen", "Bleeding LCD", "Cracked Back", "Cracked Camera", "Burns (L2-L5)",
        },
        "not_sure": {"Carrier Locked", "Not Cleared"},
    },
    "att": {
        "name": "AT&T Mobility",
        "description": "Source phones from AT&T Mobility inventory. Dispute carrier-locked units via AT&T portal — qualifies if <95% of lot unlocked by AT&T.",
        "invoice_pattern": r"at&t|at.t mobility",
        "lookback_days": 5,
        "returnable": {
            "Carrier Locked",
            # AT&T: "In the event that less than 95% of the units in an auction
            # lot are unlocked by AT&T, the order will qualify for dispute."
        },
        "not_sure": set(),
    },
    "sprint": {
        "name": "Sprint",
        "description": "Source phones from Sprint / T-Mobile auction inventory. Rules vary by cosmetic grade (NEW, A+/A/A-, B+/B/B-, C+/C/C-, D — see grade table below). Non-standard grades (CPO, DBR, DWF, N/blank) are excluded — parts-only or no grade assessment.",
        "invoice_pattern": r"sprint",
        "lookback_days": 30,
        # Fallback only — actual rules come from RMA_GRADE_RULES["Sprint"] per
        # device grade. Used if the Sprint guidelines columns are ever removed.
        "returnable": {
            "Carrier Locked",
            # Sprint: only TMO/Sprint/Metro carrier locked devices
        },
        "not_sure": set(),
    },
    "verizon": {
        "name": "Verizon Wireless",
        "description": "Source phones from Verizon Wireless inventory. Return carrier-locked and ID/MDM-locked devices via Verizon dispute — no lot threshold required.",
        "invoice_pattern": r"verizon",
        "lookback_days": 30,
        "returnable": {
            "Carrier Locked", "ID/MDM Lock",
            # Verizon: no threshold — every device qualifies
        },
        "not_sure": set(),
    },
    "superior": {
        "name": "Superior (Apple) / B Stock",
        "description": "Source phones from Superior / B Stock inventory. Return carrier-locked (>10% of order) or ID/MDM-locked (>5% of order) devices via Superior RMA.",
        "invoice_pattern": r"superior|b stock",
        "lookback_days": 5,
        "returnable": {
            "Carrier Locked", "ID/MDM Lock",
            # Superior: carrier locked if >10% of order, ID/MDM if >5%
        },
        "not_sure": set(),
    },
    "touchstone": {
        "name": "Touchstone",
        "description": "Source phones from Touchstone inventory. Return functional failures for partial credit or RTV via Touchstone — no carrier lock, cosmetic, or battery returns.",
        "invoice_pattern": r"touchstone",
        "lookback_days": 5,
        "disabled": True,
        "returnable": {
            "Camera Fail", "Buttons Fail", "Touchscreen Fail",
            "Chargeport Fail", "Parts Message", "Face ID Fail",
            "Speaker Fail", "Wifi/Bluetooth Fail",
            # Touchstone: functional failures for partial credit or RTV
            # NO Carrier Locked, NO Not Cleared, NO Low Battery, NO Cosmetic
        },
        "not_sure": set(),
    },
    "clover": {
        "name": "Clover Wireless",
        "description": "Source phones from Clover Wireless inventory. Return policy TBD — all categories pending confirmation with Jason.",
        "invoice_pattern": r"clover",
        "lookback_days": 14,  # unknown — using 14 as default
        "disabled": True,
        "returnable": set(),
        "not_sure": {
            # All rules are "NOT SURE, ASKING JASON" — flagging everything for review
            "Camera Fail", "Buttons Fail", "Touchscreen Fail",
            "Chargeport Fail", "Parts Message", "Face ID Fail",
            "Speaker Fail", "Wifi/Bluetooth Fail",
            "Low Battery", "Cosmetic Fail",
            "Carrier Locked", "ID/MDM Lock", "Not Cleared",
        },
    },
}


def load_qc_error_codes():
    """Load QC error code -> description lookup."""
    if not os.path.exists(QC_CODES_FILE):
        return {}
    df = pd.read_csv(QC_CODES_FILE, encoding="utf-8-sig")
    df = df[df["Code"].notna() & (df["Code"].str.strip() != "-- UNKNOWN --")]
    return dict(zip(df["Code"].str.strip(), df["Description"].str.strip()))


def load_invoices():
    """Load incoming invoices from the local Dropbox file (.csv or .xlsx)."""
    ext = os.path.splitext(INVOICES_FILE)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(INVOICES_FILE, dtype=str, encoding="utf-8-sig", on_bad_lines="skip")
    else:
        df = pd.read_excel(INVOICES_FILE, engine="calamine")
    # Normalize column names: replace newlines with spaces, collapse, strip
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    return df


NOTES_LOCK_PATTERN = re.compile(r"attempted to unlock|unlocking", re.IGNORECASE)


def load_detail_report(path):
    """Load an optional ICE Detail Report CSV. Returns
    {imei: {"carrier", "lock_status", "notes_locked", "retests"}}.

    Priority for deciding lock status (highest first):
      1. Carrier Lock Status column — "Unlocked" wins over "Locked" across retests;
         all-empty/N/A produces no verdict.
      2. Notes column — if any retest mentions "attempted to unlock" / "unlocking",
         the device could not be unlocked → treated as Locked. Only consulted when
         Carrier Lock Status gave no verdict.
    """
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", on_bad_lines="skip")
    df.columns = [str(c).strip() for c in df.columns]
    if "IMEI" not in df.columns or "Carrier Lock Status" not in df.columns:
        return {}
    has_carrier = "Carrier" in df.columns
    has_notes = "Notes" in df.columns
    agg = {}
    for _, row in df.iterrows():
        imei = str(row.get("IMEI") or "").strip().split(".")[0]
        if not imei or imei.lower() == "nan":
            continue
        lock_norm = str(row.get("Carrier Lock Status") or "").strip().rstrip("*").lower()
        carrier = str(row.get("Carrier") or "").strip() if has_carrier else ""
        notes = str(row.get("Notes") or "").strip() if has_notes else ""
        entry = agg.setdefault(imei, {
            "ever_unlocked": False, "ever_locked": False,
            "carrier": "", "notes_locked": False, "retests": 0,
        })
        entry["retests"] += 1
        if lock_norm == "unlocked":
            entry["ever_unlocked"] = True
        elif lock_norm == "locked":
            entry["ever_locked"] = True
        if carrier and not entry["carrier"]:
            entry["carrier"] = carrier
        if notes and NOTES_LOCK_PATTERN.search(notes):
            entry["notes_locked"] = True

    result = {}
    for imei, entry in agg.items():
        if entry["ever_unlocked"]:
            status = "Unlocked"
        elif entry["ever_locked"]:
            status = "Locked"
        else:
            status = ""  # Carrier Lock Status inconclusive — caller falls back to Notes / GSX
        result[imei] = {
            "carrier": entry["carrier"],
            "lock_status": status,
            "notes_locked": entry["notes_locked"],
            "retests": entry["retests"],
        }
    return result


def get_recent_invoices(vendor_cfg):
    """Get invoice numbers and Qty1 for a vendor within its lookback window.
    Returns (invoice_set, qty1_map) where qty1_map maps invoice# -> Qty1 (int).
    """
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    cutoff = today - timedelta(days=vendor_cfg["lookback_days"])

    df = load_invoices()
    df = df[df["Vendor"].str.contains(vendor_cfg["invoice_pattern"], case=False, na=False)]
    df["Date Received"] = pd.to_datetime(df["Date Received"], format="mixed", errors="coerce")
    df = df[(df["Date Received"] >= cutoff) & (df["Date Received"] <= today)]

    invoice_set = set(df["Invoice #"].astype(str).str.strip())

    qty1_map = {}
    for _, row in df.iterrows():
        inv = str(row["Invoice #"]).strip()
        qty1 = pd.to_numeric(row.get("QTY1"), errors="coerce")
        if pd.notna(qty1) and qty1 > 0:
            qty1_map[inv] = int(qty1)

    return invoice_set, qty1_map


def build_reasons(row, qc_lookup, returnable, not_sure, extra_pass_codes=None):
    """
    Build return reasons for a single device based on per-grade RMA rules.
    Returns (reasons, not_sure_reasons, flags):
      - reasons: issues in the grade's 'returnable' set
      - not_sure_reasons: issues in the grade's 'not_sure' set (needs testing/confirmation)
      - flags: counters used for summary stats
    """
    reasons = []
    not_sure_reasons = []
    flags = {"qc": 0, "carrier": 0, "condition": 0, "battery": 0, "id_lock": 0, "not_sure": 0}
    pass_codes = QC_PASS_CODES | (extra_pass_codes or set())

    # --- QC error codes ---
    qc_raw = row.get("QC Error Code")
    qc_code = str(qc_raw).strip() if pd.notna(qc_raw) else ""
    codes = [c.strip() for c in qc_code.split(",") if c.strip()]

    # Burn level from Physical QC LCD column — values like "L3" or "L0/L1".
    # When set to L2-L5, treat it as if the corresponding LCD-L# code were
    # present in QC Error Code (the QC code column often only carries M00/LCD-L0
    # even when Physical QC LCD records the actual burn level).
    phys_lcd_raw = row.get("Physical QC LCD")
    phys_lcd = str(phys_lcd_raw).strip().upper() if pd.notna(phys_lcd_raw) else ""
    burn_match = re.search(r'\bL([2-5])\b', phys_lcd)
    if burn_match:
        synth = f"LCD-L{burn_match.group(1)}"
        if synth not in codes:
            codes.append(synth)

    if codes:
        alarming = []
        not_sure_qc = []
        for code in codes:
            if not code or code in pass_codes:
                continue
            rma_cat = QC_CODE_TO_RMA_CATEGORY.get(code)
            desc = qc_lookup.get(code)
            # Use the QC description as the human-readable issue label (e.g.
            # "Bent Device", "Flickering / Flashing Only", "Excessive Lifted
            # Screen") rather than the umbrella RMA category — the description
            # accurately names the actual defect, while the umbrella tag would
            # mislabel things like a bent back as "Cracked Back".
            label = f"{code} ({desc})" if desc else code
            if rma_cat:
                if rma_cat in not_sure:
                    not_sure_qc.append(label)
                elif rma_cat in returnable:
                    alarming.append(label)
                # else: vendor doesn't accept this category — skip
            else:
                # Unmapped code: only flag if vendor accepts some functional failures
                if returnable & QC_FUNCTIONAL_CATEGORIES:
                    alarming.append(label)
        if alarming:
            reasons.append(f"QC: {'; '.join(alarming)}")
            flags["qc"] = 1
        if not_sure_qc:
            not_sure_reasons.append(f"QC: {'; '.join(not_sure_qc)}")
            flags["not_sure"] = 1

    # --- Carrier lock ---
    # Priority (highest first):
    #   1. Detail Report "Carrier Lock Status" column — authoritative when present
    #   2. Detail Report "Notes" column — "attempted to unlock" / "unlocking" text
    #      indicates the device could NOT be unlocked, i.e. still Locked. Consulted
    #      only when Carrier Lock Status gave no verdict. Beats Carrier GSX.
    #   3. Carrier GSX (+ brand-specific handling for Google/Pixel)
    #   4. Carrier column
    detail_lock = str(row.get("Detail Lock Status") or "").strip()
    detail_carrier = str(row.get("Detail Carrier") or "").strip()
    detail_notes_locked = bool(row.get("Detail Notes Locked"))
    carrier = str(row.get("Carrier") or "").strip()
    gsx = str(row.get("Carrier GSX") or "").strip()
    brand = str(row.get("Brand") or "").strip().lower()
    is_google = "google" in brand or "pixel" in brand

    if detail_lock:
        # Use detail report — "Locked" means carrier locked; anything else (Unlocked, N/A) is unlocked.
        if detail_lock.lower() == "locked":
            label = detail_carrier or carrier or "Unknown"
            lock_note = f"Carrier Locked: {label} (Detail Report: {detail_lock})"
            if "Carrier Locked" in returnable:
                reasons.append(lock_note)
                flags["carrier"] = 1
            elif "Carrier Locked" in not_sure:
                not_sure_reasons.append(lock_note)
                flags["not_sure"] = 1
    elif detail_notes_locked:
        # Carrier Lock Status column gave no verdict, but Notes recorded an
        # unlock attempt → device could not be unlocked → treat as Locked.
        label = detail_carrier or carrier or "Unknown"
        lock_note = f"Carrier Locked: {label} (Detail Notes: unlock attempted)"
        if "Carrier Locked" in returnable:
            reasons.append(lock_note)
            flags["carrier"] = 1
        elif "Carrier Locked" in not_sure:
            not_sure_reasons.append(lock_note)
            flags["not_sure"] = 1
    elif is_google:
        # Google/Pixel: Carrier GSX is authoritative (Locked/Unlocked); ignore Carrier column.
        if gsx.lower() == "locked":
            lock_note = f"Carrier Locked: {carrier or 'Google'} (GSX: {gsx})"
            if "Carrier Locked" in returnable:
                reasons.append(lock_note)
                flags["carrier"] = 1
            elif "Carrier Locked" in not_sure:
                not_sure_reasons.append(lock_note)
                flags["not_sure"] = 1
    elif carrier not in UNLOCKED_CARRIERS and gsx != "Unlocked":
        lock_note = f"Carrier Locked: {carrier}"
        if gsx:
            lock_note += f" (GSX: {gsx})"
        if "Carrier Locked" in returnable:
            reasons.append(lock_note)
            flags["carrier"] = 1
        elif "Carrier Locked" in not_sure:
            not_sure_reasons.append(lock_note)
            flags["not_sure"] = 1

    # --- Condition flags ---
    # The Condition column text already names the specific defect (e.g.
    # "Grade D - Back Glass Only", "Grade F - Display Failed Only"), so we
    # surface it as-is. Tagging on the umbrella RMA category would just
    # repeat "Cracked Screen" on most rows even when the real issue is
    # back glass, a bent body, or a camera lens.
    cond = str(row.get("Condition") or "").strip()
    for keyword, rma_cat in CONDITION_TO_RMA_CATEGORY.items():
        if keyword in cond.lower():
            if rma_cat in returnable:
                reasons.append(f"Condition: {cond}")
                flags["condition"] = 1
                break
            elif rma_cat in not_sure:
                not_sure_reasons.append(f"Condition: {cond}")
                flags["not_sure"] = 1
                break

    # --- Low battery ---
    if str(row.get("Physical QC Board") or "").strip() == "Battery 0-69%":
        batt_note = f"Low Battery: {row['Physical QC Board']}"
        if "Low Battery" in returnable:
            reasons.append(batt_note)
            flags["battery"] = 1
        elif "Low Battery" in not_sure:
            not_sure_reasons.append(batt_note)
            flags["not_sure"] = 1

    # --- New - Activated (from Condition field) ---
    cond_lower = cond.lower()
    if "new - activated" in cond_lower or "newly activated" in cond_lower:
        new_act_note = f"New - Activated: {cond}"
        if "New - Activated" in returnable:
            reasons.append(new_act_note)
            flags["condition"] = 1
        elif "New - Activated" in not_sure:
            not_sure_reasons.append(new_act_note)
            flags["not_sure"] = 1

    # --- ID / MDM Lock (from Condition field) ---
    if "id locked" in cond_lower or "mdm" in cond_lower or "activation lock" in cond_lower:
        id_note = f"ID/MDM Lock: {cond}"
        if "ID/MDM Lock" in returnable:
            if not any("ID/MDM Lock" in r for r in reasons):
                reasons.append(id_note)
                flags["id_lock"] = 1
        elif "ID/MDM Lock" in not_sure:
            if not any("ID/MDM Lock" in r for r in not_sure_reasons):
                not_sure_reasons.append(id_note)
                flags["not_sure"] = 1

    return reasons, not_sure_reasons, flags


def process_ice_report(ice_path, vendor_key, detail_path=None):
    """Process an uploaded ICE report against the selected vendor's rules.
    Optional detail_path points at an ICE Detail Report CSV — if provided, its
    Carrier Lock Status column overrides the default Carrier/GSX lock logic.
    """
    vendor_cfg = VENDORS[vendor_key]
    qc_lookup = load_qc_error_codes()
    detail_map = load_detail_report(detail_path) if detail_path else {}

    # Reload from disk each run so edits to the Excel file take effect immediately
    rma_grade_rules = load_rma_grade_rules()

    # Get recent invoices from Dropbox CSV
    target_invoices, qty1_map = get_recent_invoices(vendor_cfg)
    if not target_invoices:
        return None, "No recent invoices found for this vendor in the lookback window."

    # Read ICE report — tolerate reports missing the optional "Missing/Extra" column
    wanted_cols = set(ICE_COLS)
    ice = pd.read_excel(
        ice_path,
        usecols=lambda c: c in wanted_cols,
        dtype={"IMEI": str, "WID": str},
        engine="openpyxl",
    )
    ice["Vendor Invoice #"] = ice["Vendor Invoice #"].astype(str).str.strip()
    if "Pallet #" in ice.columns:
        ice["Pallet #"] = pd.to_numeric(ice["Pallet #"], errors="coerce").astype("Int64")
    ice = ice[ice["Vendor Invoice #"].isin(target_invoices)].copy()

    if ice.empty:
        return None, f"No matching devices found for {len(target_invoices)} recent invoices."

    # Drop devices flagged as Missing or Extra — they aren't sent for RMA.
    if "Missing/Extra" in ice.columns:
        me_norm = ice["Missing/Extra"].astype(str).str.strip().str.lower()
        ice = ice[~me_norm.isin({"missing", "extra"})].copy()
        if ice.empty:
            return None, "All matching devices were flagged as Missing or Extra."

    # Filter by vendor grade:
    #   - Must contain a "+" (i.e. is a graded device: A+, B+, TPS B+, DLS A+, etc.)
    #   - Must NOT be explicitly labeled for the OTHER vendor (single-program vendors)
    #     e.g. HYLA TPS keeps "B+", "TPS A+" but drops "DLS B+"
    #          HYLA DLS keeps "B+", "DLS B+" but drops "TPS A+"
    #   - Combined vendors (grade_prefix_vendors) keep ALL graded devices
    exclude_prefix = vendor_cfg.get("exclude_grade_prefix")
    grade_prefix_vendors = vendor_cfg.get("grade_prefix_vendors")
    if exclude_prefix or grade_prefix_vendors:
        vendor_cond = ice["Vendor Condition"].astype(str).str.strip()
        has_plus    = vendor_cond.str.contains(r"\+", regex=True)
        if exclude_prefix:
            not_excluded = ~vendor_cond.str.upper().str.startswith(exclude_prefix.upper())
            grade_mask   = has_plus & not_excluded
        else:
            # Combined vendor: keep all graded devices from any sub-program
            grade_mask = has_plus
        ice = ice[grade_mask].copy()
        if ice.empty:
            return None, "No graded devices found for this vendor."

    # --- Tech Notes / Notes block (HYLA only) ---
    # "Battery Door Replaced/Replacement", "Lifted Battery Door", "Lifted Door",
    # or "Lifted Back" in Tech Notes (or in Notes if it hasn't been moved to Tech
    # Notes yet) normally means the device cannot be returned via HYLA.
    # Exception: A+ and B+ grades (TPS or DLS) can still claim a lifted back —
    # those devices stay on the results and get "Lifted Back" added as a reason
    # so they end up on the RMA Claim Form.
    tech_notes_blocked_count = 0
    lifted_back_keep_mask = pd.Series(False, index=ice.index)
    if vendor_key in ("hyla", "hyla_tps", "hyla_dls"):
        lifted_pattern = r"battery door replac|lifted (?:battery door|door|back)"
        def _has_lifted(col):
            if col not in ice.columns:
                return pd.Series(False, index=ice.index)
            return (
                ice[col].astype(str).str.strip().str.lower()
                .str.contains(lifted_pattern, na=False, regex=True)
            )
        lifted_mask = _has_lifted("Tech Notes") | _has_lifted("Notes")
        if lifted_mask.any():
            grade_series = (
                ice["Vendor Condition"].astype(str).str.strip().str.upper()
                .str.split().str[-1].fillna("")
            )
            lifted_back_keep_mask = lifted_mask & grade_series.isin({"A+", "B+"})
            block_mask = lifted_mask & ~lifted_back_keep_mask
            tech_notes_blocked_count = int(block_mask.sum())
            if tech_notes_blocked_count > 0:
                ice = ice[~block_mask].copy()
                lifted_back_keep_mask = lifted_back_keep_mask.loc[ice.index]

    ice["IMEI"] = ice["IMEI"].fillna("").str.split(".").str[0]

    # Capture all WIDs/IMEIs from the full matched scope (before candidate filtering)
    all_wids  = [str(w).strip() for w in ice["WID"].fillna("")  if str(w).strip()]
    all_imeis = [str(i).strip() for i in ice["IMEI"].fillna("") if str(i).strip()]

    # Merge Detail Report lock status by IMEI (if provided)
    if detail_map:
        imei_series = ice["IMEI"].astype(str).str.strip()
        ice["Detail Lock Status"] = imei_series.map(
            lambda i: detail_map.get(i, {}).get("lock_status", "")
        )
        ice["Detail Carrier"] = imei_series.map(
            lambda i: detail_map.get(i, {}).get("carrier", "")
        )
        ice["Detail Notes Locked"] = imei_series.map(
            lambda i: detail_map.get(i, {}).get("notes_locked", False)
        )
    else:
        ice["Detail Lock Status"] = ""
        ice["Detail Carrier"] = ""
        ice["Detail Notes Locked"] = False

    # Build per-row grade-level rules and reasons
    fallback_returnable = vendor_cfg["returnable"]
    fallback_not_sure   = vendor_cfg.get("not_sure", set())

    def resolve_effective_vendor(vendor_condition):
        """For combined vendors, pick the sub-vendor based on Vendor Condition prefix."""
        if not grade_prefix_vendors:
            return vendor_cfg["name"]
        cond_upper = str(vendor_condition or "").strip().upper()
        for prefix, sub_vendor in grade_prefix_vendors.items():
            if cond_upper.startswith(prefix):
                return sub_vendor
        return vendor_cfg["name"]

    def row_build(row):
        effective_vendor = resolve_effective_vendor(row.get("Vendor Condition", ""))
        returnable, not_sure = get_grade_rules(
            effective_vendor,
            row.get("Vendor Condition", ""),
            fallback_returnable,
            fallback_not_sure,
            rma_grade_rules,
        )
        # Grade not in guidelines → suppress this device entirely
        if returnable is None:
            return [], [], {"qc": 0, "carrier": 0, "condition": 0, "battery": 0, "id_lock": 0, "not_sure": 0}
        # HYLA: ignore carrier lock when Vendor Description contains vzw/verizon
        extra_pass = None
        if vendor_key in ("hyla", "hyla_tps", "hyla_dls"):
            extra_pass = HYLA_EXTRA_PASS_CODES
            vdesc = str(row.get("Vendor Description") or "").lower()
            if "vzw" in vdesc or "verizon" in vdesc:
                returnable = returnable - {"Carrier Locked"}
                not_sure = not_sure - {"Carrier Locked"}
        return build_reasons(row, qc_lookup, returnable, not_sure, extra_pass)

    results = ice.apply(row_build, axis=1)
    ice["Reason(s)"]              = results.apply(lambda x: " | ".join(x[0]))
    ice["Not Sure / Need to Test"] = results.apply(lambda x: " | ".join(x[1]))
    flag_df = results.apply(lambda x: x[2]).apply(pd.Series)

    # HYLA A+/B+ devices with "lifted back" / "battery door" in their notes were
    # kept above; surface that here so they appear in candidates and on the form.
    if bool(lifted_back_keep_mask.any()):
        keep_aligned = lifted_back_keep_mask.reindex(ice.index, fill_value=False)
        for idx in ice.index[keep_aligned]:
            existing = ice.at[idx, "Reason(s)"]
            ice.at[idx, "Reason(s)"] = (
                f"{existing} | Lifted Back" if existing else "Lifted Back"
            )
            flag_df.at[idx, "condition"] = 1

    # Keep devices with a confirmed reason OR a "Not Sure / Need to Test" flag
    # (e.g. Display Spots is "Not sure, need to test" per the RMA sheet but we
    # still want to surface those phones so they can be reviewed.)
    candidates = ice[
        (ice["Reason(s)"] != "") | (ice["Not Sure / Need to Test"] != "")
    ].copy()

    if candidates.empty:
        return None, f"Scanned {len(ice)} devices but none flagged for return."

    # Sort by priority
    candidates["_qc"] = ~candidates["Reason(s)"].str.contains("QC:", na=False)
    candidates["_cond"] = ~candidates["Reason(s)"].str.contains("Condition:", na=False)
    candidates["_gsx"] = ~candidates["Reason(s)"].str.contains("GSX: Locked", na=False)
    candidates["_carrier"] = ~(
        candidates["Reason(s)"].str.contains("Carrier Locked:", na=False)
        & ~candidates["Reason(s)"].str.contains("GSX:", na=False)
    )
    candidates = candidates.sort_values(["_qc", "_cond", "_gsx", "_carrier"])
    candidates = candidates.drop(columns=["_qc", "_cond", "_gsx", "_carrier"])

    # Rename columns for output
    candidates = candidates.rename(columns={
        "Vendor Invoice #": "Invoice",
        "Cost Amount": "Cost",
    })

    # --- Threshold calculation ---
    rma_thresholds = load_rma_thresholds()
    rma_name = VENDOR_KEY_TO_RMA_NAME.get(vendor_key)
    vendor_threshold = rma_thresholds.get(rma_name) if rma_name else None

    # Count candidates per invoice and calculate return percentage vs Qty1
    invoice_candidate_counts = candidates.groupby("Invoice").size().to_dict()
    candidates["Invoice Qty"] = candidates["Invoice"].map(
        lambda inv: qty1_map.get(inv, 0)
    )
    candidates["Return Count"] = candidates["Invoice"].map(
        lambda inv: invoice_candidate_counts.get(inv, 0)
    )

    def calc_return_pct(row):
        qty = row["Invoice Qty"]
        return round(row["Return Count"] / qty * 100, 1) if qty > 0 else 0.0

    candidates["Return %"] = candidates.apply(calc_return_pct, axis=1)

    # Determine threshold status per row (based on invoice-level percentage)
    if vendor_threshold is not None:
        if vendor_threshold == 0.0:
            candidates["Threshold Met"] = "YES"
        else:
            candidates["Threshold Met"] = (
                candidates["Return %"] > vendor_threshold
            ).map({True: "YES", False: "NO"})
    else:
        candidates["Threshold Met"] = "NO"

    # LCD mismatch: any disagreement between Condition / QC Error Code / Physical QC LCD.
    candidates["LCD Mismatch"] = candidates.apply(
        lambda r: "YES" if lcd_mismatch(
            r.get("Condition"), r.get("QC Error Code"), r.get("Physical QC LCD"),
        ) else "NO",
        axis=1,
    )

    out_cols = [
        "WID", "Vendor", "Invoice", "Project #", "Vendor Condition", "Condition",
        "Vendor Description", "Vendor Model",
        "Brand", "Model", "Carrier", "Carrier GSX",
    ]
    if detail_map:
        out_cols += ["Detail Lock Status"]
    out_cols += [
        "QC Error Code", "Physical QC LCD", "Physical QC Board",
        "Inventory", "Inventory Bin #", "Inventory Location #",
        "Inventory Notes (Out Reason)", "Inventory By", "Inventory Date",
        "IMEI", "Cost", "Notes", "Reason(s)", "Not Sure / Need to Test",
        "Invoice Qty", "Return %", "Threshold Met", "LCD Mismatch",
    ]
    # Optional HYLA columns — included only if the ICE report had them
    for col in ("Tech Notes", "Pallet #", "Vendor Notes"):
        if col in candidates.columns:
            out_cols.append(col)
    candidates = candidates[out_cols]

    # Build stats
    stats = flag_df.loc[candidates.index].sum()
    gsx_locked    = candidates["Reason(s)"].str.contains("GSX: Locked", na=False).sum()
    carrier_no_gsx = int(stats.get("carrier", 0)) - gsx_locked
    not_sure_count = (candidates["Not Sure / Need to Test"] != "").sum()

    # Build per-grade rule breakdown for display
    grade_rules_table = {}
    if grade_prefix_vendors:
        # Combined vendor: merge sub-vendor tables, labeling grades with their prefix
        for prefix, sub_vendor_name in grade_prefix_vendors.items():
            sub_rules = rma_grade_rules.get(sub_vendor_name, {})
            for grade, grade_data in sorted(sub_rules.items()):
                labeled = f"{prefix} {grade}"
                grade_rules_table[labeled] = {
                    "returnable": sorted(cat for cat, val in grade_data.items() if val == "YES"),
                    "not_sure":   sorted(cat for cat, val in grade_data.items() if val == "NOT SURE"),
                    "no":         sorted(cat for cat, val in grade_data.items() if val == "NO"),
                }
        any_carrier = any(
            any("YES" == rma_grade_rules.get(sv, {}).get(g, {}).get("Carrier Locked")
                for g in rma_grade_rules.get(sv, {}))
            for sv in grade_prefix_vendors.values()
        )
    else:
        vendor_rules = rma_grade_rules.get(vendor_cfg["name"], {})
        if vendor_rules:
            for grade, grade_data in sorted(vendor_rules.items()):
                grade_rules_table[grade] = {
                    "returnable": sorted(cat for cat, val in grade_data.items() if val == "YES"),
                    "not_sure":   sorted(cat for cat, val in grade_data.items() if val == "NOT SURE"),
                    "no":         sorted(cat for cat, val in grade_data.items() if val == "NO"),
                }
        any_carrier = bool(
            any(
                "YES" == vendor_rules.get(g, {}).get("Carrier Locked")
                for g in vendor_rules
            )
            if vendor_rules else "Carrier Locked" in fallback_returnable
        )

    # Detail report match stats — count from candidates (return-flagged devices only)
    detail_matched = int((candidates["Detail Lock Status"] != "").sum()) if detail_map else 0
    detail_locked  = int((candidates["Detail Lock Status"].str.lower() == "locked").sum()) if detail_map else 0

    summary = {
        "total": len(candidates),
        "total_scanned": len(ice),
        "qc": int(stats.get("qc", 0)),
        "condition": int(stats.get("condition", 0)),
        "battery": int(stats.get("battery", 0)),
        "id_lock": int(stats.get("id_lock", 0)),
        "not_sure": int(not_sure_count),
        "gsx_locked": gsx_locked,
        "carrier_no_gsx": carrier_no_gsx,
        "has_carrier_check": any_carrier,
        "detail_used": bool(detail_map),
        "detail_matched": detail_matched,
        "detail_locked": detail_locked,
        "tech_notes_blocked": tech_notes_blocked_count,
        "vendor_name": vendor_cfg["name"],
        "vendor_description": vendor_cfg.get("description", ""),
        "invoices_found": len(target_invoices),
        "lookback_days": vendor_cfg["lookback_days"],
        "grade_rules_table": grade_rules_table,   # per-grade breakdown (replaces flat lists)
        "vendor_threshold": vendor_threshold,      # threshold % from RMA guidelines (None if N/A)
        "all_wids":  all_wids,
        "all_imeis": all_imeis,
    }

    return candidates, summary


# Sentinel used by the dropdown to scan every active vendor in one pass
ALL_VENDORS_KEY = "all"


def get_all_vendor_keys():
    """Vendor keys to run when the user picks 'All Vendors'.
    Skips disabled vendors and HYLA sub-views (hyla_tps/hyla_dls) — those are
    already covered by the combined `hyla` entry, so including them would
    double-count graded HYLA devices.
    """
    return [
        k for k, v in VENDORS.items()
        if not v.get("disabled") and v.get("include_in_all", True)
    ]


def process_all_vendors(ice_path, detail_path=None):
    """Run process_ice_report once per included vendor.
    Returns (per_vendor, errors) where per_vendor is a list of
    (vendor_key, candidates_df, summary) for vendors that produced results.
    """
    per_vendor = []
    errors = []
    for key in get_all_vendor_keys():
        try:
            cands, result = process_ice_report(ice_path, key, detail_path)
        except Exception as e:
            errors.append(f"{VENDORS[key]['name']}: {e}")
            continue
        if cands is None:
            errors.append(f"{VENDORS[key]['name']}: {result}")
            continue
        per_vendor.append((key, cands, result))
    return per_vendor, errors


def build_combined_summary(per_vendor, errors):
    """Aggregate per-vendor summaries into one summary dict for the template."""
    summaries = [s for _, _, s in per_vendor]
    all_wids = []
    all_imeis = []
    for s in summaries:
        all_wids.extend(s.get("all_wids", []))
        all_imeis.extend(s.get("all_imeis", []))
    return {
        "total":          int(sum(s["total"] for s in summaries)),
        "total_scanned":  int(sum(s["total_scanned"] for s in summaries)),
        "qc":             int(sum(s["qc"] for s in summaries)),
        "condition":      int(sum(s["condition"] for s in summaries)),
        "battery":        int(sum(s["battery"] for s in summaries)),
        "id_lock":        int(sum(s["id_lock"] for s in summaries)),
        "not_sure":       int(sum(s["not_sure"] for s in summaries)),
        "gsx_locked":     int(sum(s["gsx_locked"] for s in summaries)),
        "carrier_no_gsx": int(sum(s["carrier_no_gsx"] for s in summaries)),
        "has_carrier_check": any(s["has_carrier_check"] for s in summaries),
        "detail_used":    any(s["detail_used"] for s in summaries),
        "detail_matched": int(sum(s["detail_matched"] for s in summaries)),
        "detail_locked":  int(sum(s["detail_locked"] for s in summaries)),
        "tech_notes_blocked": int(sum(s.get("tech_notes_blocked", 0) for s in summaries)),
        "vendor_name":    "All Vendors (Combined)",
        "vendor_description": "Scanned across all active vendors using each vendor's own rules and lookback window.",
        "invoices_found": int(sum(s["invoices_found"] for s in summaries)),
        "lookback_days":  "varies",
        "grade_rules_table": {},
        "vendor_threshold": None,
        "all_wids":  all_wids,
        "all_imeis": all_imeis,
        "vendor_breakdown": [
            {
                "key": k,
                "name": s["vendor_name"],
                "total": s["total"],
                "lookback_days": s["lookback_days"],
                "invoices_found": s["invoices_found"],
            }
            for (k, _, s) in per_vendor
        ],
        "warnings": errors,
    }


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", vendors=VENDORS)


@app.route("/process", methods=["POST"])
def process():
    vendor_key = request.form.get("vendor")
    is_all = vendor_key == ALL_VENDORS_KEY
    if not is_all and (vendor_key not in VENDORS or VENDORS[vendor_key].get("disabled")):
        flash("Please select a valid vendor.", "error")
        return redirect(url_for("index"))

    file = request.files.get("ice_file")
    if not file or file.filename == "":
        flash("Please upload an ICE report file.", "error")
        return redirect(url_for("index"))

    filename = secure_filename(file.filename)
    if not filename.endswith((".xlsx", ".xls")):
        flash("Please upload an Excel file (.xlsx or .xls).", "error")
        return redirect(url_for("index"))

    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    file.save(filepath)

    # Optional ICE Detail Report (CSV) — used to override carrier lock status
    detail_file = request.files.get("detail_file")
    detail_path = None
    if detail_file and detail_file.filename:
        detail_name = secure_filename(detail_file.filename)
        if not detail_name.lower().endswith(".csv"):
            if os.path.exists(filepath):
                os.remove(filepath)
            flash("Detail Report must be a CSV file.", "error")
            return redirect(url_for("index"))
        detail_path = os.path.join(app.config["UPLOAD_FOLDER"], detail_name)
        detail_file.save(detail_path)

    # Build per-vendor list: one entry for single vendor, many for "all".
    try:
        if is_all:
            per_vendor, scan_errors = process_all_vendors(filepath, detail_path)
        else:
            cands, result_or_msg = process_ice_report(filepath, vendor_key, detail_path)
            if cands is None:
                per_vendor, scan_errors = [], [result_or_msg]
            else:
                per_vendor, scan_errors = [(vendor_key, cands, result_or_msg)], []
    except Exception as e:
        flash(f"Error processing file: {str(e)}", "error")
        return redirect(url_for("index"))
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
        if detail_path and os.path.exists(detail_path):
            os.remove(detail_path)

    if not per_vendor:
        msg = scan_errors[0] if scan_errors else "No return candidates found."
        if is_all and len(scan_errors) > 1:
            msg = "No return candidates found for any vendor. " + " | ".join(scan_errors)
        flash(msg, "warning")
        return redirect(url_for("index"))

    # Combine candidates and summary
    if is_all:
        candidates = pd.concat([c for _, c, _ in per_vendor], ignore_index=True)
        result = build_combined_summary(per_vendor, scan_errors)
    else:
        candidates = per_vendor[0][1]
        result = per_vendor[0][2]

    # Store results in a temp CSV for download
    csv_path = os.path.join(
        app.config["UPLOAD_FOLDER"],
        f"{vendor_key}_Return_Candidates.csv",
    )
    candidates.to_csv(csv_path, index=False, encoding="utf-8-sig")

    table_data = candidates.to_dict("records")
    columns = candidates.columns.tolist()

    # Persist each project + its phones to the DB, collect share tokens.
    # In "all" mode each project is upserted under the actual vendor that
    # produced it, so the shareable page reflects the right vendor name.
    project_links = []
    try:
        for vkey, vcands, vresult in per_vendor:
            grouped = vcands.where(pd.notnull(vcands), None).groupby("Project #", sort=False, dropna=False)
            for project_number, group in grouped:
                proj_num = "" if project_number is None else str(project_number).strip()
                if not proj_num or proj_num.lower() == "nan":
                    continue
                return_count = len(group)
                token = db.upsert_project(
                    project_number=proj_num,
                    vendor_key=vkey,
                    vendor_name=vresult["vendor_name"],
                    return_count=return_count,
                )
                db.replace_phones(proj_num, group.to_dict("records"))
                project_links.append({
                    "project_number": proj_num,
                    "token": token,
                    "return_count": return_count,
                    "url": url_for("public_project", token=token, _external=True),
                })
    except Exception as e:
        flash(f"Results saved locally, but could not write to database: {e}", "warning")

    result["project_links"] = project_links

    return render_template(
        "index.html",
        vendors=VENDORS,
        results=table_data,
        columns=columns,
        summary=result,
        selected_vendor=vendor_key,
        download_file=f"{vendor_key}_Return_Candidates.csv",
    )


def _invoices_for_vendor_cfg(df, vendor_cfg, today):
    """Filter the invoice DataFrame down to one vendor's lookback window."""
    cutoff = today - timedelta(days=vendor_cfg["lookback_days"])
    sub = df[df["Vendor"].str.contains(vendor_cfg["invoice_pattern"], case=False, na=False)].copy()
    sub["Date Received"] = pd.to_datetime(sub["Date Received"], format="mixed", errors="coerce")
    sub = sub[(sub["Date Received"] >= cutoff) & (sub["Date Received"] <= today)]
    sub = sub.sort_values("Date Received", ascending=False)
    return sub, cutoff


def _invoice_rows(df):
    has_project = "Project Number" in df.columns
    rows = []
    for _, row in df.iterrows():
        project = ""
        if has_project:
            raw_proj = row.get("Project Number")
            if pd.notna(raw_proj):
                project = str(raw_proj).strip()
                if project.lower() == "nan":
                    project = ""
        rows.append({
            "invoice": str(row["Invoice #"]).strip(),
            "date": row["Date Received"].strftime("%m/%d/%Y") if pd.notna(row["Date Received"]) else "",
            "vendor": str(row["Vendor"]).strip(),
            "project": project,
        })
    return rows


@app.route("/api/invoices/<vendor_key>")
def api_invoices(vendor_key):
    """Return recent invoices for a vendor (or all vendors) as JSON."""
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    df = load_invoices()

    if vendor_key == ALL_VENDORS_KEY:
        # Aggregate across every vendor that participates in "all" mode.
        # Each vendor's lookback window is honored; an invoice that matches
        # multiple vendor patterns is deduped by (Invoice #, Vendor).
        seen = set()
        invoices = []
        earliest_cutoff = today
        for k in get_all_vendor_keys():
            cfg = VENDORS[k]
            sub, cutoff = _invoices_for_vendor_cfg(df, cfg, today)
            if cutoff < earliest_cutoff:
                earliest_cutoff = cutoff
            for row in _invoice_rows(sub):
                dedupe_key = (row["invoice"], row["vendor"])
                if dedupe_key in seen:
                    continue
                seen.add(dedupe_key)
                invoices.append(row)
        invoices.sort(key=lambda r: r["date"], reverse=True)
        return jsonify({
            "vendor_name": "All Vendors",
            "lookback_days": "varies",
            "cutoff": earliest_cutoff.strftime("%m/%d/%Y"),
            "today": today.strftime("%m/%d/%Y"),
            "count": len(invoices),
            "invoices": invoices,
        })

    if vendor_key not in VENDORS:
        return jsonify({"error": "Unknown vendor"}), 404
    vendor_cfg = VENDORS[vendor_key]
    sub, cutoff = _invoices_for_vendor_cfg(df, vendor_cfg, today)
    invoices = _invoice_rows(sub)
    return jsonify({
        "vendor_name": vendor_cfg["name"],
        "lookback_days": vendor_cfg["lookback_days"],
        "cutoff": cutoff.strftime("%m/%d/%Y"),
        "today": today.strftime("%m/%d/%Y"),
        "count": len(invoices),
        "invoices": invoices,
    })


@app.route("/scan")
def scan():
    return render_template("scan.html")


# Vendors that support generating a vendor-facing Request Form (xlsx) from candidates
REQUEST_FORM_VENDORS = {"verizon", "hyla", "hyla_tps", "hyla_dls", "sprint"}

# Sprint / T-Mobile auction RMA form layout (one flat sheet).
# Columns A–K are supplied by the T-Mobile ASN (Advance Shipping Notice)
# manifest, joined to our return candidates by Serial No (= device IMEI).
# The RMA form is those ASN columns plus a Notes column built from the reasons.
SPRINT_ASN_COLUMNS = [
    "Auction Date", "SAP Customer Number", "Customer Organization",
    "Invoice Number", "Lot ID", "Carrier", "Auction Model", "Grade",
    "Serial No", "Master Carton ID", "Ship Date",
]
SPRINT_RMA_COLUMNS = SPRINT_ASN_COLUMNS + ["Notes"]
# Constants used as a fallback for candidates not found in the ASN manifest.
SPRINT_SAP_CUSTOMER_NUMBER = "0000780054"
SPRINT_CUSTOMER_ORG = "Mannapov LLC"
SPRINT_CARRIER = "T-MOBILE"
# Grade tokens recognized in the Vendor Condition column, most-specific first
# so "A+" is matched before "A" and "B-" before "B".
SPRINT_FORM_GRADES = ["A+", "A-", "B+", "B-", "C+", "C-", "NEW", "A", "B", "C", "D"]


def load_sprint_asn(path):
    """Load a T-Mobile ASN manifest into {serial_no: {column: value}} keyed by
    Serial No (the device IMEI). Returns {} if the file lacks a Serial No column.
    """
    df = pd.read_excel(path, dtype=str, engine="openpyxl")
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]
    if "Serial No" not in df.columns:
        return {}
    index = {}
    for _, row in df.iterrows():
        serial = str(row.get("Serial No") or "").strip().split(".")[0]
        if not serial or serial.lower() == "nan":
            continue
        rec = {}
        for col in SPRINT_ASN_COLUMNS:
            v = row.get(col)
            s = "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v).strip()
            rec[col] = "" if s.lower() == "nan" else s
        index.setdefault(serial, rec)
    return index


def _sprint_grade(vendor_condition):
    """Pull the device's cosmetic grade from the Vendor Condition string.
    Returns "" when no recognized grade is present."""
    cond = str(vendor_condition or "").strip().upper()
    if not cond:
        return ""
    last = cond.split()[-1]
    for g in SPRINT_FORM_GRADES:
        if last == g:
            return g
    for g in SPRINT_FORM_GRADES:
        if re.search(rf"(?<![\w+\-]){re.escape(g)}(?![\w+\-])", cond):
            return g
    return ""


def _sprint_note(row):
    """Build the Sprint RMA form 'Notes' value from a candidate's reasons.
    Produces concise, comma-separated tags (e.g. "Carrier Locked, Blacklisted").
    """
    reasons = str(row.get("Reason(s)") or "")
    not_sure = str(row.get("Not Sure / Need to Test") or "")
    qc = str(row.get("QC Error Code") or "").upper()
    cond = str(row.get("Condition") or "")
    qc_codes = {c.strip() for c in qc.split(",") if c.strip()}
    blob = f"{reasons} | {not_sure} | {cond}".lower()

    tags = []
    def add(t):
        if t and t not in tags:
            tags.append(t)

    if "carrier locked" in blob:
        add("Carrier Locked")
    if qc_codes & {"M32", "M33"} or "blacklist" in blob or "not cleared" in blob or "jailbroken" in blob:
        add("Blacklisted")
    if "M31" in qc_codes or "id/mdm" in blob or "id lock" in blob or "mdm" in blob or "activation lock" in blob:
        add("MDM Locked")

    # Remaining cosmetic / functional fragments. QC fragments look like
    # "M11 (Back Camera Picture/Video)" — keep the description, drop the code.
    # Condition fragments ("Grade D - Back Glass Only") are kept as-is.
    lock_tags = {"carrier locked", "blacklisted", "mdm locked"}
    bare_code = re.compile(r"^(?:M\d{2}|LCD-L\d|R\d+|FUNC\s*P\d+|[A-Z]\d{1,3})$", re.I)
    for chunk in (reasons, not_sure):
        for seg in chunk.split("|"):
            seg = seg.strip()
            if not seg or seg.lower() == "nan" or seg.lower().startswith("carrier locked"):
                continue
            seg = re.sub(r"^(QC|Condition|Low Battery|New - Activated)\s*:\s*", "", seg, flags=re.I).strip()
            for piece in seg.split(";"):
                piece = piece.strip()
                if not piece:
                    continue
                m = re.match(r"^[A-Za-z0-9\-\s]{1,14}\((.+)\)$", piece)
                if m:
                    label = m.group(1).strip()           # QC description inside parens
                else:
                    label = re.sub(r"\s*\((?:GSX|Detail)[^)]*\)", "", piece).strip()
                if not label or label.lower() in lock_tags or bare_code.match(label):
                    continue
                add(label)

    return ", ".join(tags)


def _sprint_form_values(row, asn):
    """Return the 12 RMA-form cell values for one device.
    Columns A–K come from the matched ASN row; if `asn` is None the device fell
    outside the manifest, so use the constants + ICE-derived model/grade and
    leave the auction-specific fields blank. Notes (L) is built from the reasons.
    """
    imei = str(row.get("IMEI") or "").strip().split(".")[0]
    note = _sprint_note(row)
    if asn:
        return [asn.get(col, "") for col in SPRINT_ASN_COLUMNS] + [note]

    model = str(row.get("Vendor Description") or "").strip()
    if not model or model.lower() == "nan":
        model = str(row.get("Vendor Model") or "").strip()
    if not model or model.lower() == "nan":
        model = str(row.get("Model") or "").strip()
    if model.lower() == "nan":
        model = ""
    return [
        "",                          # Auction Date
        SPRINT_SAP_CUSTOMER_NUMBER,  # SAP Customer Number
        SPRINT_CUSTOMER_ORG,         # Customer Organization
        "",                          # Invoice Number
        "",                          # Lot ID
        SPRINT_CARRIER,              # Carrier
        model,                       # Auction Model
        _sprint_grade(row.get("Vendor Condition")),  # Grade
        imei,                        # Serial No
        "",                          # Master Carton ID
        "",                          # Ship Date
        note,                        # Notes
    ]


def group_sprint_candidates_by_lot(df, asn_index=None):
    """Join candidates to the ASN by Serial No (= IMEI) and group their form
    rows by Lot ID. Each Lot ID becomes its own RMA form. Devices not found in
    the ASN have no Lot ID and are grouped under the "" key.

    Returns an ordered dict { lot_id: [value_row, ...] }.
    """
    asn_index = asn_index or {}
    groups = {}
    for _, row in df.iterrows():
        imei = str(row.get("IMEI") or "").strip().split(".")[0]
        if not imei or not imei.isdigit():
            continue
        asn = asn_index.get(imei)
        lot_id = (asn.get("Lot ID", "") if asn else "") or ""
        groups.setdefault(lot_id, []).append(_sprint_form_values(row, asn))
    return groups


def _build_sprint_workbook(value_rows):
    """Build a single Sprint RMA-form workbook from pre-computed 12-column rows.
    All cells are written as text so leading zeros (SAP #, dates) and 15-digit
    IMEIs are preserved exactly."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "RMA"
    ws.append(SPRINT_RMA_COLUMNS)
    for c in range(1, len(SPRINT_RMA_COLUMNS) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for values in value_rows:
        ws.append(values)

    for r in range(1, ws.max_row + 1):
        for c in range(1, len(SPRINT_RMA_COLUMNS) + 1):
            ws.cell(row=r, column=c).number_format = "@"

    for col_idx in range(1, len(SPRINT_RMA_COLUMNS) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=0,
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    return wb


def _reason_to_comment(reasons_str):
    """Convert a Reason(s) string into a short human-readable Additional Comments string.
    Uses the specific QC code description (e.g. "Minor Delamination") rather than the
    generic RMA category name (e.g. "Cracked Back") so comments reflect the actual issue.
    """
    if not reasons_str or str(reasons_str).strip() in ("", "nan"):
        return ""
    comments = []
    for part in str(reasons_str).split(" | "):
        part = part.strip()
        if not part:
            continue
        if part.startswith("Carrier Locked:"):
            carrier = re.sub(r'\s*\([^)]*\)', '', part.replace("Carrier Locked:", "")).strip()
            comments.append(f"{carrier} Carrier Locked")
        elif part.startswith("QC:") or part.startswith("Condition:"):
            qc_body = part.split(":", 1)[-1].strip()
            # Each entry is semicolon-separated: "CODE (Description) [Category]"
            # Use the description in parentheses; fall back to category only if missing.
            for entry in [e.strip() for e in qc_body.split(";") if e.strip()]:
                desc_match = re.search(r'\(([^)]+)\)', entry)
                if desc_match:
                    desc = desc_match.group(1).strip()
                    if desc and desc not in comments:
                        comments.append(desc)
                else:
                    cat_match = re.search(r'\[([^\]]+)\]', entry)
                    if cat_match:
                        cat = cat_match.group(1).strip()
                        if cat not in comments:
                            comments.append(cat)
                    else:
                        stripped = re.sub(r'\[[^\]]*\]', '', entry).strip()
                        if stripped and stripped not in comments:
                            comments.append(stripped)
        elif part.startswith("Low Battery:"):
            comments.append("Battery < 70%")
        elif part.startswith("ID/MDM Lock:"):
            comments.append("ID/MDM Lock")
        elif part.startswith("New - Activated:"):
            comments.append("New - Activated")
        else:
            comments.append(part)
    # deduplicate while preserving order
    seen, deduped = set(), []
    for c in comments:
        if c not in seen:
            seen.add(c)
            deduped.append(c)
    return " | ".join(deduped)


HYLA_RMA_TEMPLATE = os.path.join(BASE_DIR, "hyla_rma_template.xlsx")

# === HYLA "Reason for Return" auto-fill ====================================
# Dropdown values live in the template at G16:G108. Sourced from RMA_Codes.csv
# "Hyla Reasoning" column for QC codes; Condition column supplies the
# cosmetic-grade signal that QC codes alone can't express.

# Lower index = higher priority. Lock/identity issues outrank functional
# defects (the vendor MUST accept locks; a functional fail can be repaired);
# functional outranks cosmetic.
HYLA_REASON_PRIORITY = [
    "ACTIVATION_LOCK", "MDM LOCK", "USER LOCK", "SIM_LOCK",
    "FINANCED", "COUNTERFEIT",
    "NO_POWER", "DISPLAY", "DIGITIZER",
    "VIDEO/CAMERA", "FACE/TOUCH ID", "SPEAKER/MIC", "BTOOTH/WIFI/GPS",
    "FUNCTIONAL",
    "CRACKS", "COSMETIC", "SCRATCHES",
    "INCORRECT CARRIER", "INCORRECT MODEL", "INCORRECT CAPACITY",
]

QC_CODE_TO_HYLA_REASON = {
    # R-series
    "R1": "DISPLAY", "R2": "DISPLAY", "R3": "DISPLAY", "R4": "DISPLAY", "R9": "DISPLAY",
    "R5": "FUNCTIONAL", "R7": "FUNCTIONAL", "R8": "FUNCTIONAL",
    "R6": "COSMETIC",
    # M-series
    "M01": "NO_POWER", "M09": "NO_POWER",
    "M02": "DISPLAY",
    "M03": "DIGITIZER", "M28": "DIGITIZER",
    "M04": "FUNCTIONAL", "M05": "FUNCTIONAL", "M06": "FUNCTIONAL",
    "M07": "FUNCTIONAL", "M13": "FUNCTIONAL", "M14": "FUNCTIONAL",
    "M15": "FUNCTIONAL", "M16": "FUNCTIONAL", "M21": "FUNCTIONAL",
    "M23": "FUNCTIONAL", "M24": "FUNCTIONAL", "M25": "FUNCTIONAL",
    "M26": "FUNCTIONAL", "M27": "FUNCTIONAL", "M29": "FUNCTIONAL",
    "M33": "FUNCTIONAL", "M35": "FUNCTIONAL",
    "M08": "SPEAKER/MIC", "M12": "SPEAKER/MIC", "M19": "SPEAKER/MIC", "M30": "SPEAKER/MIC",
    "FUNC P01": "SPEAKER/MIC", "FUNC P02": "SPEAKER/MIC", "FUNC P03": "SPEAKER/MIC",
    "M11": "VIDEO/CAMERA", "M18": "VIDEO/CAMERA",
    "M17": "FACE/TOUCH ID",
    "M20": "BTOOTH/WIFI/GPS",
    "M22": "COUNTERFEIT",
    "M31": "ACTIVATION_LOCK",
    "M32": "FINANCED",
    "M34": "COSMETIC",
    # LCD-L0/L1/L2 are pass codes for HYLA (L2 is too subtle to prove in
    # photo/video); only L3-L5 are returnable burn issues.
    "LCD-L3": "DISPLAY", "LCD-L4": "DISPLAY", "LCD-L5": "DISPLAY",
    "LCD-F01": "DISPLAY", "LCD-F03": "DISPLAY", "LCD-F04": "DISPLAY", "LCD-F05": "DISPLAY",
    "LCD-F06": "DISPLAY", "LCD-F07": "DISPLAY", "LCD-F08": "DISPLAY", "LCD-F09": "DISPLAY",
    "LCD-F10": "DISPLAY", "LCD-F11": "DISPLAY", "LCD-F12": "DISPLAY", "LCD-F13": "DISPLAY",
    "LCD-F14": "DISPLAY",
    "LCD-P01": "DISPLAY", "LCD-P02": "DISPLAY", "LCD-P03": "DISPLAY", "LCD-P04": "DISPLAY",
    "LCD-PP1": "DISPLAY", "LCD-PP2": "DISPLAY", "LCD-PP3": "DISPLAY",
    "PHY-F01": "COSMETIC", "PHY-F02": "COSMETIC", "PHY-F03": "COSMETIC",
    "PHY-F04": "COSMETIC", "PHY-F05": "COSMETIC", "PHY-F09": "COSMETIC",
    "PHY-P01": "COSMETIC", "PHY-P02": "SCRATCHES", "PHY-P03": "COSMETIC",
    "PHY-P09": "COSMETIC", "PHY-P11": "COSMETIC", "PHY-P12": "VIDEO/CAMERA",
    "PHY-P13": "COSMETIC", "PHY-P14": "COSMETIC", "PHY-P15": "COSMETIC",
    "PHY-P16": "FUNCTIONAL", "PHY-P17": "COSMETIC",
}

# First match wins — keep specific keywords before generic grade fallbacks.
CONDITION_TO_HYLA_REASON = [
    ("camera lens",          "VIDEO/CAMERA"),
    ("front glass",          "CRACKS"),
    ("back glass",           "CRACKS"),
    ("back cover",           "COSMETIC"),
    ("bent",                 "COSMETIC"),
    ("bad lcd",              "DISPLAY"),
    ("display failed",       "DISPLAY"),
    ("display imperfection", "DISPLAY"),
    ("lifted screen",        "COSMETIC"),
    ("ber/scrap",            "FUNCTIONAL"),
    ("not cleared",          "COUNTERFEIT"),
    ("activation lock",      "ACTIVATION_LOCK"),
    ("id lock",              "ACTIVATION_LOCK"),
    ("mdm",                  "MDM LOCK"),
    ("new - activated",      "USER LOCK"),
    ("newly activated",      "USER LOCK"),
    ("grade f",              "DISPLAY"),
    ("grade d",              "CRACKS"),
]


# Credit Amount = Cost × multiplier, keyed by the picked HYLA reason.
# DISPLAY is split at use-site into burn (L3-L5 / "burn" in condition) vs bad LCD.
# Anything not listed falls back to FUNCTIONAL (0.45).
HYLA_CREDIT_MULTIPLIERS = {
    "ACTIVATION_LOCK": 0.8,
    "MDM LOCK": 0.8,
    "USER LOCK": 0.8,
    "SIM_LOCK": 0.25,
    "CRACKS": 0.33,
}
HYLA_DISPLAY_BURN_MULT = 0.3
HYLA_DISPLAY_LCD_MULT = 0.6
HYLA_DEFAULT_MULT = 0.45
LCD_BURN_PATTERN = re.compile(r"\bL[3-5]\b")


def _hyla_credit_amount(row, hyla_reason):
    """Compute Credit Amount = Cost × multiplier for a candidate row.

    Multiplier comes from HYLA_CREDIT_MULTIPLIERS; DISPLAY splits into burn vs
    bad LCD. Returns a rounded float, or None when Cost is missing/invalid.
    """
    cost_raw = row.get("Cost")
    try:
        cost = float(str(cost_raw).replace("$", "").replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    if cost <= 0:
        return None

    if hyla_reason == "DISPLAY":
        phys_lcd = str(row.get("Physical QC LCD") or "").strip().upper()
        cond_lower = str(row.get("Condition") or "").lower()
        is_burn = bool(LCD_BURN_PATTERN.search(phys_lcd)) or "burn" in cond_lower
        multiplier = HYLA_DISPLAY_BURN_MULT if is_burn else HYLA_DISPLAY_LCD_MULT
    else:
        multiplier = HYLA_CREDIT_MULTIPLIERS.get(hyla_reason, HYLA_DEFAULT_MULT)
    return round(cost * multiplier, 2)


def _pick_hyla_reason(row):
    """Pick the best HYLA 'Reason for Return' dropdown value for a candidate row.

    Reads QC Error Code, Physical QC LCD/Board, Reason(s), and Condition.
    Returns the highest-priority HYLA reason, or "" if nothing maps.
    """
    candidates = set()

    qc_raw = row.get("QC Error Code")
    qc_str = "" if qc_raw is None or (isinstance(qc_raw, float) and pd.isna(qc_raw)) else str(qc_raw)
    for code in (c.strip() for c in qc_str.split(",") if c.strip()):
        if code in QC_PASS_CODES or code in HYLA_EXTRA_PASS_CODES:
            continue
        mapped = QC_CODE_TO_HYLA_REASON.get(code)
        if mapped:
            candidates.add(mapped)

    phys_lcd = str(row.get("Physical QC LCD") or "").strip().upper()
    if re.search(r'\bL[3-5]\b', phys_lcd):
        candidates.add("DISPLAY")

    if str(row.get("Physical QC Board") or "").strip() == "Battery 0-69%":
        candidates.add("NO_POWER")

    reasons_lower = str(row.get("Reason(s)") or "").lower()
    if "carrier locked" in reasons_lower:
        candidates.add("SIM_LOCK")
    if "id/mdm lock" in reasons_lower or "activation lock" in reasons_lower:
        candidates.add("ACTIVATION_LOCK")
    if "lifted back" in reasons_lower or "battery door" in reasons_lower or "lifted door" in reasons_lower:
        candidates.add("COSMETIC")

    cond_lower = str(row.get("Condition") or "").lower()
    for keyword, reason in CONDITION_TO_HYLA_REASON:
        if keyword in cond_lower:
            candidates.add(reason)
            break

    if not candidates:
        return ""
    for reason in HYLA_REASON_PRIORITY:
        if reason in candidates:
            return reason
    return next(iter(candidates))


def _generate_hyla_rma_form(df, download_date=None):
    """Build a HYLA RMA Claim Form workbook from a candidates DataFrame.

    Loads the pre-formatted template so all colors, borders, merged cells,
    column widths, and data-validation dropdowns are preserved exactly.
    """
    import openpyxl

    if download_date is None:
        download_date = datetime.now()

    if not os.path.exists(HYLA_RMA_TEMPLATE):
        raise FileNotFoundError(
            "hyla_rma_template.xlsx not found in app directory. "
            "Re-generate it from the example file."
        )

    wb = openpyxl.load_workbook(HYLA_RMA_TEMPLATE)
    ws = wb["RMA Claim Form"]

    # ── Warehouse: first non-empty Vendor Notes value ────────────────────────
    warehouse = ""
    if "Vendor Notes" in df.columns:
        for v in df["Vendor Notes"]:
            s = str(v or "").strip()
            if s and s.lower() != "nan":
                warehouse = s
                break

    # ── Update live header fields ────────────────────────────────────────────
    ws["D5"] = download_date          # Date
    ws["D5"].number_format = "MM/DD/YYYY"
    ws["D8"] = warehouse              # Warehouse

    # ── Write device rows starting at row 16 ────────────────────────────────
    row_idx = 16
    for _, row in df.iterrows():
        imei = str(row.get("IMEI") or "").strip().split(".")[0]
        if not imei:
            continue

        model = str(row.get("Vendor Description") or "").strip()
        if model.lower() == "nan":
            model = ""

        # PO/Order#: prefer Pallet # column, fall back to Invoice number
        pallet = str(row.get("Pallet #") or "").strip()
        if not pallet or pallet.lower() == "nan":
            pallet = str(row.get("Invoice") or "").strip()
            if pallet.lower() == "nan":
                pallet = ""

        comment = _reason_to_comment(str(row.get("Reason(s)") or ""))

        # C = IMEI (integer, number format '0')
        ws.cell(row=row_idx, column=3).value = int(imei) if imei.isdigit() else imei
        # D = Model (text format '@' preserved from template)
        ws.cell(row=row_idx, column=4).value = model
        # E = PO/Order # (number format '0' preserved)
        ws.cell(row=row_idx, column=5).value = int(pallet) if pallet.isdigit() else pallet
        # G = Reason for Return — auto-picked from QC code / condition / locks
        hyla_reason = _pick_hyla_reason(row)
        ws.cell(row=row_idx, column=7).value = hyla_reason
        # F = Credit Amount — Cost × reason multiplier (formula in D7 sums col F)
        credit = _hyla_credit_amount(row, hyla_reason)
        if credit is not None:
            ws.cell(row=row_idx, column=6).value = credit
        # H = Additional Comments
        ws.cell(row=row_idx, column=8).value = comment

        row_idx += 1

    return wb


def classify_request_form_issue(reasons, qc_code, condition):
    """Map a candidate's reason/QC/condition to a Verizon Request Form 'Description of Issue'.
    Priority: Blacklisted > Finance Locked > MDM Locked > Carrier Locked.
    """
    def _clean(v):
        return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
    rx = _clean(reasons).lower()
    qc_u = _clean(qc_code).upper()
    cond_l = _clean(condition).lower()
    qc_codes = {c.strip() for c in qc_u.split(",") if c.strip()}
    if qc_codes & {"M32", "M33"} or "blacklist" in rx or "jailbroken" in rx or "not cleared" in rx:
        return "Blacklisted"
    if "customer exchange" in cond_l:
        return "Finance Locked"
    if "M31" in qc_codes or "id/mdm" in rx or "id lock" in rx or "mdm" in rx or "activation lock" in rx:
        return "MDM Locked"
    if "carrier locked" in rx:
        return "Carrier Locked"
    return "MDM Locked"


INVOICE_DATE_RE = re.compile(r"(\d{6})")


def invoice_to_order_date(invoice_number):
    """Extract MM/DD/YY from a Verizon-style invoice # like 'VZN031026G'. Falls back to today."""
    m = INVOICE_DATE_RE.search(str(invoice_number or ""))
    if m:
        s = m.group(1)
        return f"{s[0:2]}/{s[2:4]}/{s[4:6]}"
    return datetime.now().strftime("%m/%d/%y")


@app.route("/rmad/generate/<vendor_key>", methods=["GET", "POST"])
def rmad_generate(vendor_key):
    """Generate a vendor Request Form (.xlsx) from the latest candidates CSV for this vendor.

    Columns: Order Date | Purchase Order # | Article | IMEI* | Description of Issue.
    """
    if vendor_key not in VENDORS or vendor_key not in REQUEST_FORM_VENDORS:
        abort(404)

    csv_path = os.path.join(
        app.config["UPLOAD_FOLDER"],
        f"{vendor_key}_Return_Candidates.csv",
    )
    if not os.path.exists(csv_path):
        flash("No recent results to export — process an ICE report first.", "error")
        return redirect(url_for("index"))

    df = pd.read_csv(csv_path, dtype=str, encoding="utf-8-sig")

    # ── HYLA RMA Claim Form (one per invoice) ───────────────────────────────
    if vendor_key in ("hyla", "hyla_tps", "hyla_dls"):
        import io, zipfile as _zipfile

        date_str = datetime.now().strftime("%m%d%y")
        now = datetime.now()

        def _invoice_filename(pallet, warehouse):
            w_part = f"{warehouse} " if warehouse else ""
            return f"{w_part}({pallet}).xlsx"

        def _clean_num(v):
            s = str(v or "").strip()
            if s.lower() in ("", "nan"):
                return ""
            if s.endswith(".0"):
                s = s[:-2]
            return s

        def _pallet_for(inv_df):
            if "Pallet #" in inv_df.columns:
                for v in inv_df["Pallet #"]:
                    s = _clean_num(v)
                    if s:
                        return s
            return ""

        def _warehouse_for(inv_df):
            if "Vendor Notes" in inv_df.columns:
                for v in inv_df["Vendor Notes"]:
                    s = str(v or "").strip()
                    if s and s.lower() != "nan":
                        return s
            return ""

        invoices = [
            inv for inv in df["Invoice"].dropna().unique()
            if str(inv).strip() not in ("", "nan")
        ]

        if len(invoices) == 1:
            inv_df = df[df["Invoice"] == invoices[0]]
            warehouse = _warehouse_for(inv_df)
            pallet = _pallet_for(inv_df) or _clean_num(invoices[0])
            wb = _generate_hyla_rma_form(inv_df, download_date=now)
            fname = _invoice_filename(pallet, warehouse)
            out_path = os.path.join(app.config["UPLOAD_FOLDER"], fname)
            wb.save(out_path)
            return send_file(out_path, as_attachment=True, download_name=fname)

        # Multiple invoices → ZIP of individual forms
        zip_buf = io.BytesIO()
        with _zipfile.ZipFile(zip_buf, "w", _zipfile.ZIP_DEFLATED) as zf:
            for invoice in sorted(invoices):
                inv_df = df[df["Invoice"] == invoice]
                warehouse = _warehouse_for(inv_df)
                pallet = _pallet_for(inv_df) or _clean_num(invoice)
                wb = _generate_hyla_rma_form(inv_df, download_date=now)
                xl_buf = io.BytesIO()
                wb.save(xl_buf)
                zf.writestr(_invoice_filename(pallet, warehouse), xl_buf.getvalue())
        zip_buf.seek(0)
        vendor_label = VENDORS[vendor_key].get("name", vendor_key)
        zip_name = f"{vendor_label} RMA Claim Forms {date_str}.zip"
        return send_file(zip_buf, as_attachment=True, download_name=zip_name,
                         mimetype="application/zip")

    # ── Sprint / T-Mobile auction RMA form ──────────────────────────────────
    if vendor_key == "sprint":
        # The ASN (Advance Shipping Notice) manifest fills columns A–K, joined
        # to the candidates by Serial No (= IMEI). It's uploaded with the form.
        asn_index = {}
        asn = request.files.get("asn_file")
        if asn and asn.filename:
            asn_name = secure_filename(asn.filename)
            if not asn_name.lower().endswith((".xlsx", ".xls")):
                flash("ASN file must be .xlsx or .xls.", "error")
                return redirect(url_for("index"))
            asn_path = os.path.join(app.config["UPLOAD_FOLDER"], f"asn_{asn_name}")
            asn.save(asn_path)
            try:
                asn_index = load_sprint_asn(asn_path)
            finally:
                if os.path.exists(asn_path):
                    os.remove(asn_path)
            if not asn_index:
                flash("Couldn't read the ASN file — no 'Serial No' column found.", "error")
                return redirect(url_for("index"))

        # Each Lot ID is RMA'd separately, so emit one form per Lot ID. The
        # Lot ID doubles as the filename (T-Mobile convention, e.g.
        # "04212026SETB993660.xlsx"); a single lot downloads directly, multiple
        # lots come back zipped.
        groups = group_sprint_candidates_by_lot(df, asn_index)
        if not groups:
            flash("No devices to export — process an ICE report first.", "error")
            return redirect(url_for("index"))

        def _lot_filename(lot_id):
            if lot_id:
                return f"{secure_filename(lot_id)}.xlsx"
            return f"Sprint RMA {datetime.now():%m%d%y} (no Lot ID).xlsx"

        if len(groups) == 1:
            lot_id, rows = next(iter(groups.items()))
            wb = _build_sprint_workbook(rows)
            fname = _lot_filename(lot_id)
            out_path = os.path.join(app.config["UPLOAD_FOLDER"], fname)
            wb.save(out_path)
            return send_file(out_path, as_attachment=True, download_name=fname)

        # Multiple Lot IDs → ZIP of one form each
        import io as _io
        import zipfile as _zipfile

        zip_buf = _io.BytesIO()
        with _zipfile.ZipFile(zip_buf, "w", _zipfile.ZIP_DEFLATED) as zf:
            for lot_id, rows in sorted(groups.items()):
                wb = _build_sprint_workbook(rows)
                xl_buf = _io.BytesIO()
                wb.save(xl_buf)
                zf.writestr(_lot_filename(lot_id), xl_buf.getvalue())
        zip_buf.seek(0)
        zip_name = f"Sprint RMA Forms {datetime.now():%m%d%y}.zip"
        return send_file(zip_buf, as_attachment=True, download_name=zip_name,
                         mimetype="application/zip")

    # ── Verizon (and other vendors) ──────────────────────────────────────────
    from openpyxl import Workbook  # local import: only needed for this route

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Order Date", "Purchase Order #", "Article", "IMEI*", "Description of Issue"])

    for _, row in df.iterrows():
        invoice = str(row.get("Invoice") or "").strip()
        imei = str(row.get("IMEI") or "").strip().split(".")[0]
        if not imei or not imei.isdigit():
            continue
        article = str(row.get("Vendor Description") or "").strip()
        issue = classify_request_form_issue(
            row.get("Reason(s)"), row.get("QC Error Code"), row.get("Condition"),
        )
        ws.append([invoice_to_order_date(invoice), invoice, article, int(imei), issue])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).number_format = "mm-dd-yy"
        ws.cell(row=r, column=4).number_format = "0"

    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            val = cell.value
            if val is None:
                continue
            if r > 1 and col_idx == 1:
                text = str(val.strftime("%m-%d-%y")) if hasattr(val, "strftime") else str(val)
            else:
                text = str(val)
            if len(text) > max_len:
                max_len = len(text)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    date_str = datetime.now().strftime("%m%d%y")
    vendor_short = VENDORS[vendor_key]["name"].split()[0]  # "Verizon" from "Verizon Wireless"
    fname = f"{date_str} {vendor_short} Request Form.xlsx"
    out_path = os.path.join(app.config["UPLOAD_FOLDER"], fname)
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name=fname)


@app.route("/rmad/upload", methods=["POST"])
def rmad_upload():
    """Parse an uploaded Verizon-style Request Form and return its IMEIs.
    The form's 4th column (header 'IMEI*') contains the IMEIs.
    """
    file = request.files.get("rmad_file")
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "No file uploaded."}), 400
    fname = secure_filename(file.filename)
    if not fname.lower().endswith((".xlsx", ".xls")):
        return jsonify({"ok": False, "error": "File must be .xlsx or .xls"}), 400

    path = os.path.join(app.config["UPLOAD_FOLDER"], f"rmad_{fname}")
    file.save(path)
    try:
        if os.path.getsize(path) == 0:
            return jsonify({"ok": False, "error": "File is empty (0 bytes). If it's in Dropbox, open it once so Dropbox downloads it before uploading."}), 400
        try:
            df = pd.read_excel(path, header=0, dtype=str, engine="openpyxl")
        except Exception as exc:
            msg = str(exc)
            if "not a zip" in msg.lower():
                msg = "File isn't a valid .xlsx (likely a Dropbox cloud-only placeholder — open it once to materialize it locally first)."
            return jsonify({"ok": False, "error": msg}), 400
        if df.shape[1] < 4:
            return jsonify({"ok": False, "error": "Expected at least 4 columns in the form."}), 400
        imeis, seen = [], set()
        for raw in df.iloc[:, 3]:
            if pd.isna(raw):
                continue
            s = str(raw).strip().split(".")[0]
            if not s or not s.isdigit():
                continue
            if s in seen:
                continue
            seen.add(s)
            imeis.append(s)
        return jsonify({"ok": True, "imeis": imeis, "count": len(imeis)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        if os.path.exists(path):
            os.remove(path)


@app.route("/download/<filename>")
def download(filename):
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], secure_filename(filename))
    if not os.path.exists(filepath):
        flash("Download file not found. Please re-process.", "error")
        return redirect(url_for("index"))
    return send_file(filepath, as_attachment=True, download_name=filename)


# ---------------------------------------------------------------------------
# Admin auth
# ---------------------------------------------------------------------------
@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        pw = request.form.get("password", "")
        if not ADMIN_PASSWORD or ADMIN_PASSWORD == "CHANGE_ME_BEFORE_DEPLOY":
            flash("Admin password not set on the server. Edit .env.", "error")
        elif hmac.compare_digest(pw, ADMIN_PASSWORD):
            session["is_admin"] = True
            session.permanent = True
            flash("Logged in as admin.", "success")
            next_url = request.args.get("next") or url_for("index")
            return redirect(next_url)
        else:
            flash("Wrong password.", "error")
    return render_template("admin_login.html")


@app.route("/admin/logout", methods=["POST"])
def admin_logout():
    session.pop("is_admin", None)
    flash("Logged out.", "success")
    return redirect(request.referrer or url_for("index"))


# ---------------------------------------------------------------------------
# Public project page + admin photo management
# ---------------------------------------------------------------------------
@app.route("/project/<token>")
def public_project(token):
    project = db.get_project_by_token(token)
    if not project:
        abort(404)
    phones = db.get_phones_for_project(project["project_number"])
    photos = db.list_photos(project["project_number"])
    return render_template(
        "project.html", project=project, phones=phones, photos=photos, token=token,
    )


@app.route("/project/<token>/photo", methods=["POST"])
@require_admin
def add_project_photo(token):
    project = db.get_project_by_token(token)
    if not project:
        abort(404)
    photo_url = request.form.get("photo_url", "").strip()
    label = request.form.get("label", "").strip()
    if not photo_url:
        flash("Photo URL is required.", "error")
    elif not (photo_url.startswith("http://") or photo_url.startswith("https://")):
        flash("Photo URL must start with http:// or https://", "error")
    else:
        db.add_photo(project["project_number"], photo_url, label)
        flash("Photo added.", "success")
    return redirect(url_for("public_project", token=token))


@app.route("/project/<token>/photo/<int:photo_id>/delete", methods=["POST"])
@require_admin
def delete_project_photo(token, photo_id):
    project = db.get_project_by_token(token)
    if not project:
        abort(404)
    if db.delete_photo(photo_id, project["project_number"]):
        flash("Photo removed.", "success")
    else:
        flash("Photo not found.", "error")
    return redirect(url_for("public_project", token=token))


# ---------------------------------------------------------------------------
# Scanner photo uploads
# ---------------------------------------------------------------------------
PHOTOS_DIR = os.path.join(app.config["UPLOAD_FOLDER"], "photos")
ALLOWED_PHOTO_EXT = {"jpg", "jpeg", "png", "webp"}


@app.route("/project/<token>/photo/upload", methods=["POST"])
def upload_project_photo(token):
    """Upload a scanned photo and attach to a project. Admin only.
    Returns JSON so the scanner UI can show success/error per capture.
    """
    if not session.get("is_admin"):
        return jsonify({"ok": False, "error": "not_admin",
                        "message": "Log in as admin to save photos."}), 403
    project = db.get_project_by_token(token)
    if not project:
        return jsonify({"ok": False, "error": "not_found"}), 404

    file = request.files.get("photo")
    if not file or not file.filename:
        return jsonify({"ok": False, "error": "no_file"}), 400
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "jpg"
    if ext not in ALLOWED_PHOTO_EXT:
        return jsonify({"ok": False, "error": "bad_ext"}), 400

    os.makedirs(PHOTOS_DIR, exist_ok=True)
    fname = f"{secure_filename(token)}_{int(datetime.now().timestamp() * 1000)}.{ext}"
    save_path = os.path.join(PHOTOS_DIR, fname)
    file.save(save_path)

    label = (request.form.get("label") or "").strip()
    photo_url = url_for("serve_photo", filename=fname, _external=True)
    photo_id = db.add_photo(project["project_number"], photo_url, label)
    return jsonify({"ok": True, "photo_id": photo_id, "url": photo_url})


@app.route("/photos/<path:filename>")
def serve_photo(filename):
    """Serve a scanned photo from the uploads/photos directory."""
    safe = secure_filename(filename)
    if not safe or safe != filename:
        abort(404)
    return send_file(os.path.join(PHOTOS_DIR, safe))


if __name__ == "__main__":
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    os.makedirs(PHOTOS_DIR, exist_ok=True)
    # Debug mode is OFF by default. Enable locally with FLASK_DEBUG=1.
    # NEVER run with debug enabled in production — it exposes the Werkzeug
    # interactive debugger, which allows arbitrary remote code execution.
    debug = os.environ.get("FLASK_DEBUG", "").lower() in ("1", "true", "yes")
    host = os.environ.get("FLASK_HOST", "127.0.0.1")
    port = int(os.environ.get("FLASK_PORT", "5000"))
    app.run(debug=debug, host=host, port=port)
